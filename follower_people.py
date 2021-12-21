#   Tweetbot follower
#   by Tomislav Rozman
#   Usage:
#   1. Put several keywords in Tweets.xlsx, worksheet Keywords, B Column
#   2. Run the program
#   3. This followerbot will search for twitter accounts which have <keywords> in their profile 
#   and it will follow them

from openpyxl import load_workbook
from openpyxl.styles.styleable import NumberFormatDescriptor

from colorama import init, deinit # color text
from colorama import Fore, Back, Style

import tweepy
import datetime
import time
import traceback

#read twitter keys from the config file (tweerconfig.txt)
print("Reading config file with twitter keys...")
def get_pair(line): #transform string line to 2 vars: key & value
    key, sep, value = line.strip().partition("=")
    return key, value

with open("tweetconfig.txt") as config_file:    
    config_keys = dict(get_pair(line) for line in config_file) #read line from the file, convert it to key,value and insert it into dict

print("Done.")


# Your app's API/consumer key and secret can be found under the Consumer Keys
# section of the Keys and Tokens tab of your app, under the
# Twitter Developer Portal Projects & Apps page at
# https://developer.twitter.com/en/portal/projects-and-apps
consumer_key = str(config_keys["consumer_key"])
consumer_secret = str(config_keys["consumer_secret"])

# Your account's (the app owner's account's) access token and secret for your
# app can be found under the Authentication Tokens section of the
# Keys and Tokens tab of your app, under the
# Twitter Developer Portal Projects & Apps page at
# https://developer.twitter.com/en/portal/projects-and-apps
access_token = str(config_keys["access_token"])
access_token_secret = str(config_keys["access_token_secret"])

auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)

print("Authenticating to twitter again...")
# Authenticate Twitter
try:
    TweetAPI = tweepy.API(auth,wait_on_rate_limit=False)
    
    # If the authentication was successful, this should print the
    # screen name / username of the account
    print("Twitter username:"+TweetAPI.verify_credentials().screen_name)
    print("Auth. successfull!")
except:
    print("Auth. was not successfull!")
    exit()
#end try auth

init() #init colorama for coloured text

XLSfilename="Tweets.xlsx"
Worksheetname="Keywords"

print("Opening "+XLSfilename+" and getting keywords...")
wb = load_workbook(filename = XLSfilename)
ws = wb[Worksheetname] #worksheet name

keywords=[]
row=tuple

#helper vars and counters
no_of_new_follows=0
no_of_fails=0

keywords=ws.cell(2,1).value.split(",")
print("Keywords from excel:",str(keywords))
print("Search tweets with keyword ")

#search for tweets
#for keyword in keywords:
#    print(Fore.BLUE+"*** Keyword: "+keyword)    
#    for tweet in TweetAPI.search_tweets(q=keyword, lang="en", count=10):
#        print(Fore.YELLOW+tweet.user.name+"(@"+tweet.user.screen_name+")"+":"+Fore.WHITE+tweet.text+Fore.CYAN+" at "+str(tweet.created_at))
#        print("------------------------------")

print("Search tweet accounts with keyword ")
for keyword in keywords:
    print(Fore.BLUE+"*** Keyword: "+keyword)    
    for tweet_user in TweetAPI.search_users(q=keyword, count=100):
            
        print(Fore.WHITE+"Trying to follow @"+tweet_user.screen_name+" ("+tweet_user.name+")")
        
        for attempt in range(3): #retry if hit limit
            try:
                friendship=TweetAPI.get_friendship(source_screen_name="BICERO_Ltd",target_screen_name=tweet_user.screen_name)
            except Exception as e:
                print("Error getting friends, now waiting fot 16 mins. Attempt "+str(attempt))
                print(e)
                #exit()
                time.sleep(16*60) #wait for 15 mins to overcome the rate limit
            else:
                print("Attempt was successful, continuing after 6 secs...")
                time.sleep(6)
                break
        else:
            #we failed all attempts to save, exiting
            print("All attempts failed, exiting...")
            exit()
        #for - retry

        if friendship[0].following == False:
            print("BICERO_Ltd"+" is not following "+tweet_user.screen_name)
            for follow_attempt in range(3):#retry following 3 times, wait after each retry
                try:
                    TweetAPI.create_friendship(user_id=tweet_user.id)
                    time.sleep(6) #limit the number of requests
                except tweepy.HTTPException as e:
                    print("Error code:"+str(e.api_codes))
                    if e.api_codes==[160]: #follow request already sent
                        print("Follow request already sent, skipping! ")
                        print(e)
                        time.sleep(6)
                        break

                    if e.api_codes==[161]: #unable to follow more people
                        print(Fore.RED+"Follow NOT successfull, waiting for 24 hrs from "+datetime.datetime.now())
                        print(e)
                        no_of_fails+=1
                        time.sleep(24*60*60) #wait for 24 hrs
                    
                        
                else: #try except else
                    print(Fore.GREEN+"Follow successful.")
                    no_of_new_follows+=1
                    break
            else: #for
                #we failed all attempts to follow, exiting
                print("All attempts to follow failed, exiting...")
                exit()
            #for retry    
        else:
            print(Fore.LIGHTRED_EX +"BICERO_Ltd"+" is ALREADY following "+tweet_user.screen_name)
            no_of_fails+=1

        print("------------------------------")

#the main loop: repeat forewer
#Every 12 hrs read keywords from excel & follow the users


#jump back bc. while  
print("")
print(Fore.WHITE+"All done.")
print(Fore.RED+"Fails:"+str(no_of_fails))
print(Fore.GREEN+"New follows:"+str(no_of_new_follows))  
deinit() #stop coloured output