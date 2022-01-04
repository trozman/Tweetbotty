#   Tweetbot xls2tw
#   by Tomislav Rozman
#   Usage:
#   1. Put your twitter keys to 'tweetconfig.txt'  
#   2. Put some tweets to 'Tweets.xlsx', worksheet 'Tweets', B Column
#   3. Run the program: 'python xls2tw.py'
#   4. This tweetbot will run forewer: it will read tweets in sequence from excel tile, 
#   publish it to Twitter, pausing for X hours. If it finds the empty row in the 
#   spreadsheet(no more tweets), it will start from the beginning. You can add rows 
#   with tweets while this script is running.
#   If you interrupt it (ctrl-c) and re-run it, it will start from the last published tweet.
#

from openpyxl import load_workbook
from openpyxl.styles.styleable import NumberFormatDescriptor

from colorama import init, deinit, Fore, Back, Style # color text
from colorama import Fore, Back, Style

import tweepy
import datetime
import time
import logging


def find_cursor(ws):
    #Restore the cursor location from excel (Where is the * ?)
    RowCursor=2 #from the 2nd row on:
    while not ws.cell(RowCursor,1).value=="*": #repeat until the first column contains text
        print("Value of A col.:"+str(ws.cell(RowCursor,1).value))
        print("Cursor not found at "+str(RowCursor)+", increasing.")
        RowCursor=RowCursor+1
        if RowCursor==10000: #hard limit: 10 000 tweets. If cursor not found, exit.
            exit()
    return RowCursor #return the position of *



#read twitter keys from the config file (tweerconfig.txt)
print("Reading config file with twitter keys...")
def get_pair(line): #transform string line to 2 vars: key & value
    key, sep, value = line.strip().partition("=")
    return key, value

with open("tweetconfig.txt") as config_file:    
    config_keys = dict(get_pair(line) for line in config_file) #read line from the file, convert it to key,value and insert it into dict

print("Done.")

#configure logging - requires Python 3.9!
logging.basicConfig(filename='xls2tw.log', encoding='utf-8', level=logging.DEBUG)

# Your app's API/consumer key and secret can be found under the Consumer Keys
# section of the Keys and Tokens tab of your app, under the
# Twitter Developer Portal Projects & Apps page at
# https://developer.twitter.com/en/portal/projects-and-apps
consumer_key = str(config_keys["consumer_key"])
consumer_secret = str(config_keys["consumer_secret"])

# Your account's (the app owner's account's) access token and secret for your
# app can be found under the Authentication Tokens section of the
# Keys and Tokens tab of your app, under the
# Twitter Developer Portal Projects & Apps page at
# https://developer.twitter.com/en/portal/projects-and-apps
access_token = str(config_keys["access_token"])
access_token_secret = str(config_keys["access_token_secret"])

auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)

waithrs=1
waitmins=1

init() #init colorama for coloured text

XLSfilename="Tweets.xlsx"
Worksheetname="Tweets"

print("Opening "+XLSfilename+" and searching for cursor (last twitted row)...")
 
wb = load_workbook(filename = XLSfilename)
ws = wb[Worksheetname] #worksheet name

TweetText=str
NumberOfTweetsinXLS=0
row=tuple

RowCursor=2 #skip the top row in Excel (header)

wb = load_workbook(filename = XLSfilename)
ws = wb[Worksheetname] #worksheet name

RowCursor=find_cursor(ws) #call function - find * in xls
        
print("Rowcursor is at:"+str(RowCursor))

#the main loop: repeat forewer
#Read tweets from excel until empty row is found then repeat from the start
while True:
    print("Authenticating to twitter(again)...")
    # Authenticate Twitter
    try:
        TweetAPI = tweepy.API(auth)
    
        # If the authentication was successful, this should print the
        # screen name / username of the account
        print("Twitter username:"+TweetAPI.verify_credentials().screen_name)
        print("Auth. successfull!")
        logging.info("Auth. successfull"+" at:"+str(datetime.datetime.now()))
    except Exception as auth_e:
        print("Auth. was NOT successfull!")
        logging.info("Auth. was NOT successfull"+" at:"+str(datetime.datetime.now()))
        print(str(auth_e))
        exit()
    #end try auth
    
    wb = load_workbook(filename = XLSfilename)
    ws = wb[Worksheetname] #worksheet name
            
    #run through all rows and return tuples for each row - return tweet text
    
    while ws.cell(RowCursor,2).value: #repeat until the cell (column 2) contains text
        TweetText=str(ws.cell(RowCursor,2).value) # get tweet text from the spreadsheet
        tweet_time=datetime.datetime.now()
        print("Tweeting #"+str(RowCursor)+":"+TweetText+" /at "+str(tweet_time))
        try:
            TweetAPI.update_status(TweetText) #tweet
            print("Posting to Twitter OK!")
            logging.info("Posting to TW ok"+" at:"+str(datetime.datetime.now()))
            #store cursor location to excel     
            ws.cell(RowCursor,1).value="*"   #write cursor to excel
            ws.cell(RowCursor,5).value=tweet_time
            ws.cell(RowCursor-1,1).value=" "  #delete prev cursor

        except tweepy.HTTPException as e1: #handle twitter error codes
            print("Error code:"+str(e1.api_codes))
            print(e1)
            if e1.api_codes==[187]: #duplicate tweet
                print("Duplicate tweet, skipping! ")
                logging.info("Duplicate tweet, skipping"+" at:"+str(datetime.datetime.now()))
                #store cursor location to excel     
                print("Cursor is now at " + str(RowCursor))
                
                ws.cell(RowCursor,5).value=tweet_time
                ws.cell(RowCursor,1).value=" "  #delete prev cursor

                
                ws.cell(RowCursor+1,1).value="*"   #set cursor to next tweet
                time.sleep(6)
                

        except Exception as e:
            print("Posting to Twitter failed! "+str(e))
            logging.info("Posting to TW failed"+" at:"+str(datetime.datetime.now()))
            exit()
        #else: #if no exceptions

        
        
        #save excel file with new cursor location
        for save_attempt in range(10):#retry saving 10 times, wait 10 secs after each retry
            try:
                wb.save(filename=XLSfilename) #save updated excel file (with cursor and post time) 
            except:
                print("Error occured while saving excel file, retrying in 10 secs... "+str(save_attempt)+"/10")
                logging.info("Error occured while saving excel file, retrying in 10 secs..."+" at:"+str(datetime.datetime.now()))
                time.sleep(10) 
            else:
                print("Saving excel & new cursor successfully, continuing...")
                logging.info("Saving excel & new cursor successfully, continuing..."+" at:"+str(datetime.datetime.now()))
                break
        else:
            #we failed all attempts to save, exiting
            print("All attempts saving excel failed, exiting...")
            logging.info("All attempts saving excel failed, exiting..."+" at:"+str(datetime.datetime.now()))
                
            exit()

        print("Waiting for 12 hours until next tweet...")  
        logging.info("Waiting for 12 hours until next tweet..."+" at:"+str(datetime.datetime.now()))
             
       
        time.sleep(12*60*60) #sleep for 12 hrs - does not work if computer goes to standby - ensure it's always on
        #time.sleep(5) #uncomment for testing purposes, if you want quicker turnaround

        now=datetime.datetime.now()
        print("Slept for "+str((now-tweet_time)/60)+" mins.")

        RowCursor=RowCursor+1

        #reload excel - in case a new tweet was added to the excel while sleeping
        wb = load_workbook(filename = XLSfilename)
        ws = wb[Worksheetname] #worksheet name  
       
    #end while - iterating
    # remove last cursor at last tweet

    ws.cell(RowCursor,5).value=datetime.datetime.now()
    ws.cell(RowCursor-1,1).value=" "  #delete prev cursor

    #save excel file with new cursor location
    for save_attempt in range(10):#retry saving 10 times, wait 10 secs after each retry
        try:
            wb.save(filename=XLSfilename) #save updated excel file (with cursor and post time) 
        except:
            print("Error occured while saving excel file, retrying in 10 secs... "+str(save_attempt)+"/10")
            logging.info("Error occured while saving excel file, retrying in 10 secs..."+" at:"+str(datetime.datetime.now()))
         
            time.sleep(10) 
        else:
            print("Saving excel & new cursor successfully, continuing...")
            logging.info("Saving excel & new cursor successfully, continuing..."+" at:"+str(datetime.datetime.now()))
         
            break
    else:
         #we failed all attempts to save, exiting
        print("All attempts saving excel failed, exiting...")
        logging.info("All attempts saving excel failed, exiting..."+" at:"+str(datetime.datetime.now()))
         
        exit()

    RowCursor=2 #reset row cursor, because we run out of tweets, therefore we'll begin at the 2nd row
    print("*****Out of tweets, starting again from the beginning...")
    logging.info("*****Out of tweets, starting again from the beginning..."+" at:"+str(datetime.datetime.now()))
         
    #jump back bc. while    
    deinit() #stop coloured output